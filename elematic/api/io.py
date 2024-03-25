# import sys
import re
import openpyxl
from . import MatML_api


def import_xml(input_file):
    """Reads MatML information from an XML file. Currently does not
    validate the file.

    Args:
        inputFile (str): filename location of XML data.

    Returns:
        elematic.api.MatML_api.MatML_doc: the root element of the MatML
        data
    """
    doc = MatML_api.parsexml_(input_file)
    rootNode = doc.getroot()
    rootObject = MatML_api.MatML_Doc.factory()
    rootObject.build(rootNode)
    doc = None
    return rootObject


def import_granta_maptis_mmpds(input_file: str) -> MatML_api.MatML_Doc:
    """Converts a XLSX workbook from MAPTIS GRANTA MMPDS to a MatML_Doc.

    Args:
        input_file (str): filename location of the XLSX workbook

    Returns:
        MatML_api.MatML_Doc: result of the conversion
    """

    # Initial MatML_Doc with metadata and a material card with empty bulk
    # details
    library = MatML_api.MatML_Doc()
    library.Metadata = MatML_api.Metadata()
    metadata = library.Metadata
    library.add_Material(MatML_api.Material())
    material = library.Material[0]
    material.BulkDetails = MatML_api.BulkDetails()
    bulk_details = material.BulkDetails

    # Load workbook from provided filepath
    workbook = openpyxl.load_workbook(input_file)
    # MAPTIS template uses the "Export Lookup" sheet to list all data and their
    # locations.
    lookup_sheet = workbook["Export Lookup"]
    # Iterate through each line item in export lookup sheet. Excel sheets index
    # from 1, so we start at row 2 since first row has headers
    for row in lookup_sheet.iter_rows(min_row=2, max_row=100):
        # Get information about the target data
        target_worksheet_name = row[4].value
        target_range = row[6].value
        target_unit_range = row[2].value

        # Dependent data are specified when the target worksheet name ends with
        # a tilde
        if target_worksheet_name.endswith("~"):
            # The GRANTA MMPDS export template on MAPTIS appends a "1_In" to
            # dependent data sheets
            target_worksheet_name += "1_In"
            # Data which are dependent on multiple parameters (e.g., tensile
            # data depends on temperature AND time), there are usually multiple
            # series (e.g., _01, _02, _03, etc. up to _20 or so). For now only
            # first series is imported. TODO: allow for multiple series.
            target_range += "_01"

        # Check if the target worksheet actually exists. If not, skip the row.
        try:
            target_worksheet = workbook[target_worksheet_name]
        except KeyError:
            continue

        # Read target data from XLSX
        cell_values = _get_cell_value(
            target_range_name=target_range, target_worksheet=target_worksheet
        )

        # For now, we skip footnotes and source figures. TODO: create
        # MatML_api.Note and MatML_api.DataSourceDetails elements for these.
        if "FOOTNOTE" in target_range:
            continue
        if "SOURCEFIGURE" in target_range:
            continue

        # Check to make sure we actually found data at the target destination.
        if cell_values is None:
            continue
        if len(cell_values) == 0:
            continue

        # Otherwise, convert the XLSX data to MatML data.
        match target_range:
            case "MI_RECORDNAME":
                continue
            case "MI_SPECIFICATION":
                bulk_details.add_Specification(
                    MatML_api.Specification(valueOf_=cell_values[0])
                )
            case "MI_SHORTNAME":
                continue
            case "MI_RECORDGUID":
                continue
            case "MI_COMMONNAME":
                bulk_details.Name = MatML_api.Name(valueOf_=cell_values[0])
            case "MI_CONDITION":
                bulk_details.add_ProcessingDetails(
                    Name=MatML_api.Name(valueOf_=cell_values[0])
                )
            case "MI_THICKNESS":
                # Initialize Form, Geometry, Size, and Dimensions if they have
                # not been set, yet.
                if bulk_details.Form is None:
                    bulk_details.Form = MatML_api.Form()
                if bulk_details.Form.Geometry is None:
                    bulk_details.Form.Geometry = MatML_api.Geometry()
                    bulk_details.Form.Geometry.Shape = ""
                # If Dimensions already exist, then we add another entry to it.
                if bulk_details.Form.Geometry.Dimensions is None:
                    bulk_details.Form.Geometry.Dimensions = ""
                else:
                    bulk_details.Form.Geometry.Dimensions += ", "
                # Form the dimension string. There is a "TRUE" or "FALSE" entry
                # next to each dimension value indicating if it is inclusive
                # ("TRUE") or exclusive ("FALSE").
                dimension_string = str(cell_values[0])
                dimension_string += (
                    " ≤ thickness" if cell_values[1] == "TRUE" else " < thickness"
                )
                if len(cell_values) == 4:
                    dimension_string += " ≤ " if cell_values[3] == "TRUE" else " < "
                    dimension_string += str(cell_values[2])
                # Add unit string (which is found in the second column of the
                # Export Lookup sheet)
                dimension_string += " " + _get_cell_value_from_destination(
                    row[2].value, workbook=workbook
                )
                bulk_details.Form.Geometry.Dimensions += dimension_string
            case "MI_AREA":
                # Initialize Form, Geometry, Size, and Dimensions if they have
                # not been set, yet.
                if bulk_details.Form is None:
                    bulk_details.Form = MatML_api.Form()
                if bulk_details.Form.Geometry is None:
                    bulk_details.Form.Geometry = MatML_api.Geometry()
                    bulk_details.Form.Geometry.Shape = ""
                # If Dimensions already exist, then we add another entry to it.
                if bulk_details.Form.Geometry.Dimensions is None:
                    bulk_details.Form.Geometry.Dimensions = ""
                else:
                    bulk_details.Form.Geometry.Dimensions += ", "
                # Form the dimension string. There is a "TRUE" or "FALSE" entry
                # next to each dimension value indicating if it is inclusive
                # ("TRUE") or exclusive ("FALSE").
                dimension_string = str(cell_values[0])
                dimension_string += " ≤ area" if cell_values[1] == "TRUE" else " < area"
                if len(cell_values) == 4:
                    dimension_string += " ≤ " if cell_values[3] == "TRUE" else " < "
                    dimension_string += str(cell_values[2])
                # Add unit string (which is found in the second column of the
                # Export Lookup sheet)
                dimension_string += " " + _get_cell_value_from_destination(
                    row[2].value, workbook=workbook
                )
                bulk_details.Form.Geometry.Dimensions += dimension_string
            case "MI_WIDTH":
                # Initialize Form, Geometry, Size, and Dimensions if they have
                # not been set, yet.
                if bulk_details.Form is None:
                    bulk_details.Form = MatML_api.Form()
                if bulk_details.Form.Geometry is None:
                    bulk_details.Form.Geometry = MatML_api.Geometry()
                    bulk_details.Form.Geometry.Shape = ""
                # If Dimensions already exist, then we add another entry to it.
                if bulk_details.Form.Geometry.Dimensions is None:
                    bulk_details.Form.Geometry.Dimensions = ""
                else:
                    bulk_details.Form.Geometry.Dimensions += ", "
                # Form the dimension string. There is a "TRUE" or "FALSE" entry
                # next to each dimension value indicating if it is inclusive
                # ("TRUE") or exclusive ("FALSE").
                dimension_string = str(cell_values[0])
                dimension_string += (
                    " ≤ width" if cell_values[1] == "TRUE" else " < width"
                )
                if len(cell_values) == 4:
                    dimension_string += " ≤ " if cell_values[3] == "TRUE" else " < "
                    dimension_string += str(cell_values[2])
                # Add unit string (which is found in the second column of the
                # Export Lookup sheet)
                dimension_string += " " + _get_cell_value_from_destination(
                    row[2].value, workbook=workbook
                )
                bulk_details.Form.Geometry.Dimensions += dimension_string
            case "MI_AVAILABLEFORMS":
                if bulk_details.Form is None:
                    bulk_details.Form = MatML_api.Form()
                bulk_details.Form.Description = MatML_api.Name(value_of=cell_values[0])
            case "MI_STATISTICALBASIS":
                # It would probably be wiser to put the statistical basis as a
                # Qualifier for relevant property data in accordance with
                # MMPDS Chapter 9 customary statistical basis... TODO
                if bulk_details.Notes is None:
                    bulk_details.Notes = "Statistical Basis: " + cell_values[0]
                else:
                    bulk_details.Notes += "\nStatistical Basis: " + cell_values[0]
            case "MI_SOURCEFIGURE":
                # Similar to note above... probably better to include this
                # information as a "data source" rather than just in the Notes
                # but this works well enough for now. TODO
                if bulk_details.Notes is None:
                    bulk_details.Notes = "Source Figure: " + cell_values[0]
                else:
                    bulk_details.Notes += "\nSource Figure: " + cell_values[0]
            case _:
                # Otherwise... treat the data as though it is PropertyData.
                data_value, data_format = _convert_list_to_string(cell_values)
                if data_value is None:
                    continue

                # Add the property data to the bulk details.
                bulk_details.add_PropertyData(
                    MatML_api.PropertyData(property=target_range)
                )
                property = bulk_details.get_PropertyData(property=target_range)[0]
                property.Data = MatML_api.DataType(
                    valueOf_=data_value,
                    format=data_format,
                )

                # Add a corresponding PropertyDetails entry to the metadata.
                metadata.add_PropertyDetails(MatML_api.PropertyDetails(id=target_range))
                property_details = metadata.get_PropertyDetails(id=target_range)
                property_name = row[0].value
                property_details.Name = MatML_api.Name(valueOf_=property_name)
                # Sometimes the units column is a destination, other times it is
                # a unit string directly.
                property_unit_string = row[2].value

                if property_unit_string is None:
                    property_details.Unitless = MatML_api.Unitless()
                else:
                    if "=" in property_unit_string:
                        property_unit_string = _get_cell_value_from_destination(
                            destination=property_unit_string, workbook=workbook
                        )
                    property_details.Units = _form_units(
                        unit_string=property_unit_string
                    )

                # Now we check for parameter values.
                for col in range(17, 24):
                    # Check each column for ranges.
                    parameter_range_name = row[col].value
                    if parameter_range_name is None:
                        break
                    if target_worksheet_name.endswith("_In"):
                        # The data are multi-series so we append "_01".
                        parameter_range_name += "_01"

                    # Get the parameter data.
                    parameter_cell_value = _get_cell_value(
                        target_range_name=parameter_range_name,
                        target_worksheet=target_worksheet,
                    )
                    # Process the data and obtain its format.
                    parameter_data, parameter_format = _convert_list_to_string(
                        parameter_cell_value
                    )
                    if parameter_data is None:
                        break
                    # Add the ParameterValue.
                    property.add_ParameterValue(
                        MatML_api.ParameterValue(
                            parameter=parameter_range_name,
                            Data=MatML_api.DataType(
                                valueOf_=parameter_data, format=parameter_format
                            ),
                        )
                    )

                    # Add the ParameterDetails.
                    parameter_name, parameter_units = _get_parameter_metadata_info(
                        parameter=parameter_range_name, workbook=workbook
                    )
                    if parameter_units is None:
                        metadata.add_ParameterDetails(
                            MatML_api.ParameterDetails(
                                id=parameter_range_name,
                                Name=MatML_api.Name(valueOf_=parameter_name),
                                Unitless=MatML_api.Unitless(),
                            )
                        )
                    else:
                        metadata.add_ParameterDetails(
                            MatML_api.ParameterDetails(
                                id=parameter_range_name,
                                Name=MatML_api.Name(valueOf_=parameter_name),
                                Units=_form_units(parameter_units),
                            )
                        )
    return library


def _get_cell_value(target_range_name: str, target_worksheet) -> list:
    """Given a range name and worksheet, returns the value(s) of the cell(s)
    in that range.

    Args:
        target_range_name (_type_): _description_
        target_worksheet (_type_): _description_

    Returns:
        list: values
    """
    cells = []

    # Start by getting the cell objects corresponding with the data range. Make
    # sure the range name is in the defined names of the worksheet.
    if target_range_name in target_worksheet.defined_names:
        # If it is a named range (i.e., in defined_names), then we need to get
        # the destination corresponding to that range first.
        definition = target_worksheet.defined_names[target_range_name]
        destinations = definition.destinations
        for title, cell_range in destinations:
            if title == target_worksheet.title:
                cells = target_worksheet[cell_range]
    # If a colon exists in the range name, it is a direct range (e.g., =B2:B20)
    elif ":" in target_range_name:
        # In this case, we directly access the cell objects.
        cells = target_worksheet[target_range_name]

    # If cell objects were found, then we collect the values at those cells.
    if cells:
        # If cells is a tuple, that means the range covers multiple cells.
        if isinstance(cells, tuple):
            values = []
            # Iterate through each row and column covered by the range.
            for row in cells:
                row_values = []
                for cell in row:
                    # If the cell is empty, then there won't be anymore data
                    # after this column in this row, so we break.
                    if cell.value is None:
                        break
                    # Otherwise, append the cell value.
                    row_values.append(cell.value)
                if row_values:
                    # If the row is empty, there won't be anymore data after
                    # this row, so we break.
                    values.append(row_values)
                else:
                    break
            values = [item for sublist in values for item in sublist]
            return values
        # If cells is not a tuple, that means the range corresponds with a
        # single cell.
        else:
            return [cells.value]
    # If no cells were found, then we return an empty list.
    return []


def _get_cell_value_from_destination(
    destination,
    worksheet: openpyxl.worksheet.worksheet.Worksheet = None,
    workbook: openpyxl.Workbook = None,
) -> openpyxl.worksheet.cell_range.CellRange:
    """Given a destination and either worksheet or workbook, gets the value of
    the cell corresponding to that destination.

    Note: currently only matches absolute destinations (i.e., destination must
    include a sheetname! in it.)

    Args:
        destination (str): string in the form of either "=<sheetname>!$<column>$<row>"
        worksheet (openpyxl.worksheet.worksheet.Worksheet, optional)
        workbook (openpyxl.Workbook, optional)

    Raises:
        ValueError: if no sheetname is provided, then the worksheet must be supplied
        ValueError: at least one of worksheet or workbook must be supplied
        ValueError: worksheet must be supplied when sheetname is not in destination

    Returns:
        openpyxl.worksheet.cell_range.CellRange: cell_range value
    """
    if "!" not in destination:
        if worksheet == None:
            raise ValueError(
                "worksheet must be provided when destination does not include sheet name (has no '!' in it)"
            )
    # Regex to match groups of <sheet_name>!$<column>$<row>
    match = re.match(r"=(\w+)!\$(\w+)\$(\d+)", destination)
    if match:
        sheet_name, column, row = match.groups()
        if workbook is not None:
            return workbook[sheet_name][column + row].value
        elif worksheet is not None:
            return workbook[column + row].value
        else:
            raise ValueError("either worksheet or workbook must be provided.")
    else:
        # Regex to match groups of $<column>$<row>
        match = re.match(r"=\$(\w+)\$(\d+)", destination)
        if match:
            column, row = match.groups()
            if worksheet is None:
                raise ValueError("Worksheet must be provided.")
            return worksheet[column + row].value


def _convert_list_to_string(data: list) -> list[str, str]:
    """Takes a list of data and returns a comma-separated string representation
    according to the MatML 3.1 XSD data formats.

    - Note: currently assumes that entire list is of same data format... TODO:
    consider checking if mixed formats are provided.
    - Note: returns a float even if the data format is "integer"... this mostly
    has to do with the fact that Python treats the values as floats when mapping
    them to a str.

    Args:
        data (list): _description_

    Returns:
        list[str,str]: the comma-separated list of data and its format
    """
    # If the provided list is None, or if it has entries but the first
    # entry is None, then we exit the function
    if not data or data[0] is None:
        return None, None
    # If the first entry is a str, then we assume they all are strings
    if isinstance(data[0], str):
        return [",".join(data), "string"]
    # Otherwise treat float and int types the same
    elif isinstance(data[0], float | int):
        return [",".join(map(str, data)), "float"]


def _form_units(unit_string: str) -> MatML_api.Units:
    units = MatML_api.Units()
    numerators = unit_string.split("/")[0]
    denominators = unit_string.split("/")[1:]

    # This regex splits on all "." characters that are not surrounded on both
    # sides by digits. So something like "ksi.in^0.5" would be split into "ksi"
    # and "in^0.5", but not "ksi", "in^0" and "5". The loop iterates through
    # each of these groups.
    for unit_and_power in re.split(r"(?<![0-9])\.(?![0-9])", numerators):

        # Now we try to split each group by the "^". This regex splits on "^" if
        # there is a letter to the left and a number to the right. So a string
        # like "in^0.5" would split into "in" and "0.5". But a string like
        # "10^6 psi" would not split at all. Sometimes no power is specified,
        # in which case the re.split returns a list of length 1 and we set the
        # power by default to "1" since this is a unit in the numerator.
        split_name_and_power = re.split(r"(?<![0-9])\^(?![^0-9])", unit_and_power)
        if len(split_name_and_power) == 1:
            unit_power = "1"
        else:
            unit_power = split_name_and_power[1]
        unit_name = split_name_and_power[0]
        units.add_Unit(
            MatML_api.Unit(
                power=unit_power,
                description=unit_name + "^" + unit_power,
                Name=str(unit_name),
            )
        )

    # Now we repeat the same procedure, except for each denominator unit.
    # Unfortunately, the formatting isn't entirely consistent here. For the
    # coefficient of thermal expansion units, the string would be like
    # "in/in/°F". But for thermal conductivity, the unit string is
    # "BTU/hr.ft.°F".
    for denominator in denominators:
        for unit_and_power in re.split(r"(?<![0-9])\.(?![0-9])", denominator):
            split_name_and_power = re.split(r"(?<![0-9])\^(?![^0-9])", unit_and_power)
            if len(split_name_and_power) == 1:
                unit_power = "1"
            else:
                unit_power = split_name_and_power[1]
            unit_power = "-" + unit_power
            unit_name = split_name_and_power[0]
            units.add_Unit(
                MatML_api.Unit(
                    power=unit_power,
                    description=unit_name + "^" + unit_power,
                    Name=str(unit_name),
                )
            )

    return units


def _get_parameter_metadata_info(
    parameter: str,
    workbook: openpyxl.Workbook,
    parameter_lookup_sheetname: str = "Parameter Lookup",
) -> list[str, str]:
    """Obtains the name and the unit string of the parameter value.

    Args:
        parameter (str): the data range name of the parameter
        workbook (openpyxl.Workbook): the workbook to search
        parameter_lookup_sheetname (str): the name of the parameter lookup sheet
        that contains the corresponding data. Defaults to "Parameter Lookup".

    Returns:
        list[str,str]: the name and unit string of the parameter.
    """
    if parameter.endswith("_01"):
        parameter = parameter.strip("_01")
    for row in workbook[parameter_lookup_sheetname].iter_rows(min_row=2):
        if row[1].value == parameter:
            return [row[0].value, row[2].value]
    return [None, None]
